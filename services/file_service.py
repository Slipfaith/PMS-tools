# services/file_service.py - ПОЛНАЯ ВЕРСИЯ С ПОДДЕРЖКОЙ SDLXLIFF

import sqlite3
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, Optional, List
import logging
import re

logger = logging.getLogger(__name__)


class FileService:
    """Сервис для работы с файлами. Вся логика анализа файлов здесь."""

    def __init__(self):
        self.supported_formats = {
            '.sdltm': 'SDL Trados Memory',
            '.xlsx': 'Excel Workbook',
            '.xls': 'Excel Workbook',
            '.tmx': 'TMX Memory',
            '.xml': 'XML/Termbase',
            '.mtf': 'MultiTerm Format',
            '.tbx': 'TBX Termbase',
            '.sdlxliff': 'SDL XLIFF'
        }

    def get_file_info(self, filepath: Path) -> Dict[str, any]:
        """Получает всю информацию о файле для GUI"""
        try:
            info = {
                'name': filepath.name,
                'size_mb': filepath.stat().st_size / (1024 * 1024),
                'format': self.get_format_name(filepath),
                'format_icon': self.get_format_icon(filepath),
                'extra_info': '',
                'is_supported': self.is_supported(filepath)
            }

            # Дополнительная информация в зависимости от формата
            if filepath.suffix.lower() == '.sdltm':
                info['extra_info'] = self._get_sdltm_info(filepath)
            elif filepath.suffix.lower() in ['.xlsx', '.xls']:
                info['extra_info'] = self._get_excel_info(filepath)
            elif filepath.suffix.lower() == '.sdlxliff':
                info['extra_info'] = self._get_sdlxliff_info(filepath)

            return info

        except Exception as e:
            logger.warning(f"Error analyzing file {filepath}: {e}")
            return {
                'name': filepath.name,
                'size_mb': 0,
                'format': 'Unknown',
                'format_icon': '📄',
                'extra_info': f'Ошибка: {e}',
                'is_supported': False
            }

    def is_supported(self, filepath: Path) -> bool:
        """Проверяет, поддерживается ли формат"""
        return filepath.suffix.lower() in self.supported_formats

    def get_format_name(self, filepath: Path) -> str:
        """Возвращает название формата"""
        suffix = filepath.suffix.lower()
        return self.supported_formats.get(suffix, 'Unknown Format')

    def get_format_icon(self, filepath: Path) -> str:
        """Возвращает иконку для формата"""
        suffix = filepath.suffix.lower()
        icons = {
            '.sdltm': '🗄️',
            '.xlsx': '📊',
            '.xls': '📊',
            '.tmx': '🔄',
            '.xml': '📋',
            '.mtf': '📖',
            '.tbx': '📖',
            '.sdlxliff': '✂️'
        }
        return icons.get(suffix, '📄')

    def detect_files_format(self, filepaths: List[str]) -> tuple[str, List[str]]:
        """Определяет формат файлов для drop area"""
        valid_files = []
        detected_formats = set()
        sdlxliff_parts = []

        for filepath in filepaths:
            path = Path(filepath)
            if path.exists() and path.is_file() and self.is_supported(path):
                valid_files.append(filepath)

                # Особая обработка для SDLXLIFF частей
                if path.suffix.lower() == '.sdlxliff' and self.is_sdlxliff_part(path):
                    sdlxliff_parts.append(path)
                    detected_formats.add("SDL XLIFF части")
                else:
                    detected_formats.add(self.get_format_name(path))

        if not valid_files:
            return "Неподдерживаемый формат", []

        # Если есть части SDLXLIFF, проверяем, все ли части присутствуют
        if sdlxliff_parts:
            # Группируем части по базовому имени
            parts_groups = {}
            for part_path in sdlxliff_parts:
                part_info = self.get_sdlxliff_part_info(part_path)
                if part_info:
                    base_name = re.sub(r'\.\d+of\d+\.sdlxliff, '', str(part_path))
                    if base_name not in parts_groups:
                        parts_groups[base_name] = {
                            'total': part_info['total'],
                            'found': []
                        }
                    parts_groups[base_name]['found'].append(part_info['part'])

                    # Формируем информацию о частях
                    parts_info = []
                    for base_name, info in parts_groups.items():
                        found_count = len(info['found'])
                    total_count = info['total']
                    if found_count == total_count:
                        parts_info.append(f"✅ Все {total_count} частей")
                    else:
                        parts_info.append(f"⚠️ {found_count} из {total_count} частей")

                    format_name = f"SDL XLIFF части ({', '.join(parts_info)})"
                    elif len(detected_formats) == 1:
                    format_name = list(detected_formats)[0]
                    else:
                    format_name = f"Смешанные форматы ({len(detected_formats)})"

        return format_name, valid_files

    def auto_detect_languages(self, filepath: Path) -> Optional[Dict[str, str]]:
        """Автоопределение языков из SDLTM файла"""
        if filepath.suffix.lower() != '.sdltm':
            return None

        try:
            with sqlite3.connect(str(filepath)) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT source_segment, target_segment FROM translation_units LIMIT 10")

                src_lang = "unknown"
                tgt_lang = "unknown"

                for src_xml, tgt_xml in cursor.fetchall():
                    try:
                        # Парсим source
                        if src_lang == "unknown":
                            root = ET.fromstring(src_xml)
                            lang_elem = root.find(".//CultureName")
                            if lang_elem is not None and lang_elem.text:
                                src_lang = self._normalize_language(lang_elem.text)

                        # Парсим target
                        if tgt_lang == "unknown":
                            root = ET.fromstring(tgt_xml)
                            lang_elem = root.find(".//CultureName")
                            if lang_elem is not None and lang_elem.text:
                                tgt_lang = self._normalize_language(lang_elem.text)

                        # Если нашли оба языка, прекращаем поиск
                        if src_lang != "unknown" and tgt_lang != "unknown":
                            break

                    except Exception:
                        continue

                if src_lang != "unknown" or tgt_lang != "unknown":
                    return {"source": src_lang, "target": tgt_lang}

        except Exception as e:
            logger.warning(f"Error detecting languages from {filepath}: {e}")

        return None

    def _get_sdltm_info(self, filepath: Path) -> str:
        """Получает информацию об SDLTM файле"""
        try:
            with sqlite3.connect(str(filepath)) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM translation_units")
                count = cursor.fetchone()[0]
                return f"{count:,} сегментов"
        except Exception:
            return "Недоступно"

    def _get_excel_info(self, filepath: Path) -> str:
        """Получает информацию об Excel файле"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(filepath), read_only=True)

            sheets_count = len(wb.sheetnames)

            # Подсчитываем примерное количество строк
            total_rows = 0
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if sheet.max_row > 1:  # Исключаем пустые листы
                    total_rows += sheet.max_row - 1  # Минус заголовок

            wb.close()

            if total_rows > 0:
                return f"{sheets_count} листов, ~{total_rows:,} строк"
            else:
                return f"{sheets_count} листов"

        except Exception as e:
            logger.warning(f"Error analyzing Excel file {filepath}: {e}")
            return "Требует настройки"

    def _get_sdlxliff_info(self, filepath: Path) -> str:
        """Получает информацию о SDLXLIFF файле"""
        try:
            # Проверяем, является ли файл частью
            match = re.search(r'\.(\d+)of(\d+)\.sdlxliff, str(filepath), re.IGNORECASE)
            if match:
                part = match.group(1)
            total = match.group(2)
            return f"Часть {part} из {total}"
            else:
            # Пытаемся получить базовую информацию без полного анализа
            # (чтобы не создавать циклическую зависимость)
            try:
                # Быстрая проверка - является ли файл XML
                with open(filepath, 'rb') as f:
                    # Читаем первые 1KB для быстрой проверки
                    header = f.read(1024)
                    if b'<xliff' in header or b'trans-unit' in header:
                        # Оцениваем размер для примерного количества сегментов
                        file_size_mb = filepath.stat().st_size / (1024 * 1024)
                        estimated_segments = int(file_size_mb * 100)  # Примерная оценка
                        return f"~{estimated_segments:,} сегментов"
                    else:
                        return "Требует анализа"
            except:
                return "Требует анализа"

    except Exception as e:
    logger.warning(f"Error analyzing SDLXLIFF file {filepath}: {e}")
    return "Недоступно"


def is_sdlxliff_part(self, filepath: Path) -> bool:
    """Проверяет, является ли файл частью разделенного SDLXLIFF"""
    pattern = r'\.\d+of\d+\.sdlxliff
    return bool(re.search(pattern, str(filepath), re.IGNORECASE))


def get_sdlxliff_part_info(self, filepath: Path) -> Optional[Dict[str, int]]:
    """Извлекает информацию о части SDLXLIFF файла"""
    match = re.search(r'\.(\d+)of(\d+)\.sdlxliff, str(filepath), re.IGNORECASE)
    if match:
        return {
            'part': int(match.group(1)),
            'total': int(match.group(2))
        }
    return None


def _normalize_language(self, lang_code: str) -> str:
    """Нормализует языковой код"""
    if not lang_code:
        return "unknown"

    # Стандартные замены
    lang_map = {
        "en": "en-US", "de": "de-DE", "fr": "fr-FR", "it": "it-IT",
        "es": "es-ES", "pt": "pt-PT", "ru": "ru-RU", "ja": "ja-JP",
        "ko": "ko-KR", "zh": "zh-CN", "pl": "pl-PL", "tr": "tr-TR"
    }

    code = lang_code.lower().replace("_", "-")

    # Если уже полный код
    if "-" in code and len(code) == 5:
        return code

    # Добавляем регион по умолчанию
    return lang_map.get(code, f"{code}-XX")