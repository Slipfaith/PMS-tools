# core/converters/excel_converter.py - СУПЕР ОПТИМИЗИРОВАННАЯ ВЕРСИЯ

import openpyxl
from pathlib import Path
from typing import Iterator, Tuple, Dict, List, Optional, Set
import logging
import time

from ..base import (
    StreamingConverter, ConversionResult, ConversionOptions, ConversionStatus,
    ExcelAnalysis, SheetInfo, ColumnInfo, ColumnType, ExcelConversionSettings,
    TranslationSegment, ExcelStructureError, ValidationError, ConversionError
)
from ..formats.tmx_format import TmxWriter

logger = logging.getLogger(__name__)


class ExcelConverter(StreamingConverter):
    """Супер быстрый конвертер Excel в TMX"""

    def __init__(self):
        super().__init__()
        self.supported_formats = {'.xlsx', '.xls'}

    def can_handle(self, filepath: Path) -> bool:
        """Проверяет, может ли конвертер обработать файл"""
        return filepath.suffix.lower() in self.supported_formats

    def validate(self, filepath: Path) -> bool:
        """Быстрая валидация Excel файла"""
        if not filepath.exists():
            raise ValidationError(f"File not found: {filepath}")

        if not filepath.is_file():
            raise ValidationError(f"Not a file: {filepath}")

        try:
            # Минимальная проверка - только открываем и закрываем
            wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
            has_sheets = bool(wb.sheetnames)
            wb.close()

            if not has_sheets:
                raise ValidationError(f"No sheets found in Excel file: {filepath}")

            return True

        except Exception as e:
            raise ValidationError(f"Invalid Excel file: {e}")

    def analyze_excel_structure(self, filepath: Path) -> ExcelAnalysis:
        """Супер быстрый анализ структуры"""
        logger.info(f"Fast analyzing Excel: {filepath.name}")

        analysis = ExcelAnalysis(file_path=filepath)

        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)

        for sheet_name in wb.sheetnames[:10]:  # Максимум 10 листов
            ws = wb[sheet_name]

            # Быстрый анализ
            sheet_info = SheetInfo(name=sheet_name)

            # Читаем только первую строку для заголовков
            headers = []
            for col in range(1, min(ws.max_column + 1, 21)):  # Максимум 20 колонок
                value = ws.cell(1, col).value
                if value:
                    headers.append(ColumnInfo(
                        index=col - 1,
                        name=str(value),
                        column_type=ColumnType.TEXT
                    ))

            sheet_info.columns = headers
            sheet_info.data_rows = ws.max_row - 1 if ws.max_row > 1 else 0

            analysis.sheets.append(sheet_info)

        wb.close()
        return analysis

    def convert_excel_to_tmx(self, filepath: Path, settings: ExcelConversionSettings,
                             options: ConversionOptions) -> ConversionResult:
        """СУПЕР БЫСТРАЯ конвертация Excel в TMX с поддержкой многих листов"""
        start_time = time.time()

        try:
            self._update_progress(0, "Открываем Excel...", options)

            # Используем iter_rows для максимальной скорости
            wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)

            all_segments = []
            total_sheets = len(settings.selected_sheets)

            logger.info(f"Processing {total_sheets} sheets: {settings.selected_sheets}")

            for sheet_idx, sheet_name in enumerate(settings.selected_sheets):
                if self._should_stop(options):
                    wb.close()
                    return self._create_cancelled_result()

                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                    continue

                ws = wb[sheet_name]
                column_mapping = settings.get_sheet_column_mapping(sheet_name)

                if not column_mapping:
                    logger.warning(f"No column mapping for sheet '{sheet_name}'")
                    continue

                # Находим нужные колонки
                source_col_idx = None
                target_col_idx = None
                comment_col_indices = []

                for col_info in column_mapping.values():
                    if col_info.final_type == ColumnType.TEXT:
                        if col_info.final_language == settings.source_language:
                            source_col_idx = col_info.index
                        elif col_info.final_language == settings.target_language:
                            target_col_idx = col_info.index
                    elif col_info.final_type == ColumnType.COMMENT:
                        comment_col_indices.append(col_info.index)

                # Автоопределение если не указано
                if source_col_idx is None or target_col_idx is None:
                    text_cols = [col for col in column_mapping.values()
                                 if col.final_type == ColumnType.TEXT]
                    if len(text_cols) >= 2:
                        source_col_idx = text_cols[0].index
                        target_col_idx = text_cols[1].index

                if source_col_idx is None or target_col_idx is None:
                    logger.error(f"Sheet '{sheet_name}': could not determine source/target columns")
                    continue

                logger.info(f"Sheet '{sheet_name}': source_col={source_col_idx}, target_col={target_col_idx}")

                # СУПЕР БЫСТРОЕ чтение с iter_rows
                row_count = 0
                sheet_segments = []

                # Обновляем прогресс для текущего листа
                sheet_progress_base = int((sheet_idx / total_sheets) * 70)
                self._update_progress(
                    sheet_progress_base,
                    f"Обрабатываем лист '{sheet_name}'...",
                    options
                )

                # Читаем ВСЕ данные ОДНИМ проходом
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_count += 1

                    # Быстрое извлечение значений
                    try:
                        source_text = str(row[source_col_idx] or '').strip()
                        target_text = str(row[target_col_idx] or '').strip()

                        if not source_text and not target_text:
                            continue

                        segment = TranslationSegment(
                            source_text=source_text,
                            target_text=target_text,
                            source_lang=settings.source_language,
                            target_lang=settings.target_language
                        )

                        # Комментарии только если нужно
                        if settings.include_comments and comment_col_indices:
                            for comment_idx in comment_col_indices:
                                if comment_idx < len(row) and row[comment_idx]:
                                    segment.add_comment(str(row[comment_idx]))

                        sheet_segments.append(segment)

                    except Exception as e:
                        logger.debug(f"Error processing row {row_count}: {e}")
                        continue

                    # Обновляем прогресс каждые 1000 строк
                    if row_count % 1000 == 0:
                        sheet_progress = sheet_progress_base + int((row_count / ws.max_row) * (70 / total_sheets))
                        self._update_progress(
                            sheet_progress,
                            f"Лист '{sheet_name}': {row_count} строк",
                            options
                        )

                all_segments.extend(sheet_segments)
                logger.info(f"Sheet '{sheet_name}': extracted {len(sheet_segments)} segments from {row_count} rows")

            wb.close()

            # Быстрая фильтрация дубликатов
            self._update_progress(80, f"Обработка {len(all_segments)} сегментов...", options)

            if settings.skip_empty_segments or True:  # Всегда фильтруем
                seen = set()
                unique_segments = []

                for seg in all_segments:
                    if not seg.source_text or not seg.target_text:
                        continue

                    key = (seg.source_text.lower(), seg.target_text.lower())
                    if key not in seen:
                        seen.add(key)
                        unique_segments.append(seg)
            else:
                unique_segments = all_segments

            if not unique_segments:
                return ConversionResult(
                    success=False,
                    output_files=[],
                    stats={
                        "total": len(all_segments),
                        "exported": 0,
                        "processed_sheets": len([s for s in settings.selected_sheets if s in wb.sheetnames])
                    },
                    errors=["No valid segments found after filtering"],
                    status=ConversionStatus.FAILED
                )

            # Быстрая запись TMX
            self._update_progress(90, f"Сохраняем {len(unique_segments)} уникальных сегментов...", options)

            output_path = filepath.with_suffix('.tmx')
            tmx_segments = [
                (seg.source_text, seg.target_text, seg.source_lang, seg.target_lang)
                for seg in unique_segments
            ]

            TmxWriter.write(output_path, tmx_segments,
                            settings.source_language, settings.target_language)

            elapsed = time.time() - start_time
            stats = {
                "total_segments": len(all_segments),
                "exported_segments": len(unique_segments),
                "duplicates_removed": len(all_segments) - len(unique_segments),
                "processed_sheets": len([s for s in settings.selected_sheets if s in wb.sheetnames]),
                "conversion_time": elapsed,
                "segments_per_second": int(len(all_segments) / elapsed) if elapsed > 0 else 0,
                "source_language": settings.source_language,
                "target_language": settings.target_language
            }

            self._update_progress(100, f"Готово за {elapsed:.1f} сек! {len(unique_segments)} сегментов", options)

            return ConversionResult(
                success=True,
                output_files=[output_path],
                stats=stats,
                status=ConversionStatus.COMPLETED
            )

        except Exception as e:
            logger.exception(f"Error converting Excel: {e}")
            return ConversionResult(
                success=False,
                output_files=[],
                stats={"error": str(e)},
                errors=[str(e)],
                status=ConversionStatus.FAILED
            )

    def _create_cancelled_result(self) -> ConversionResult:
        """Создает результат для отмененной конвертации"""
        return ConversionResult(
            success=False,
            output_files=[],
            stats={"cancelled": True},
            errors=["Conversion cancelled by user"],
            status=ConversionStatus.CANCELLED
        )

    def convert(self, filepath: Path, options: ConversionOptions) -> ConversionResult:
        """Основной метод конвертации"""
        try:
            analysis = self.analyze_excel_structure(filepath)

            if not analysis.sheets:
                raise ConversionError("No sheets found in Excel file")

            # Автоматические настройки
            settings = ExcelConversionSettings(
                source_language="ru-RU",
                target_language="en-US",
                selected_sheets=[sheet.name for sheet in analysis.sheets if sheet.is_selected]
            )

            for sheet in analysis.sheets:
                if sheet.is_selected:
                    column_mapping = {}
                    for col in sheet.columns:
                        column_mapping[col.index] = col
                    settings.column_mappings[sheet.name] = column_mapping

            return self.convert_excel_to_tmx(filepath, settings, options)

        except Exception as e:
            logger.exception(f"Error in Excel conversion: {e}")
            return ConversionResult(
                success=False,
                output_files=[],
                stats={"error": str(e)},
                errors=[str(e)],
                status=ConversionStatus.FAILED
            )

    def convert_streaming(self, filepath: Path, options: ConversionOptions) -> Iterator[Tuple[str, str, str, str]]:
        """Потоковая конвертация"""
        return iter([])

    def get_progress_steps(self, filepath: Path) -> int:
        """Возвращает примерное количество шагов"""
        return 100