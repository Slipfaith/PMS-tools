# core/formats/xlsx_format.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from typing import List, Tuple
import logging

logger = logging.getLogger(__name__)


class XlsxWriter:
    """Полный XLSX writer с форматированием"""

    @classmethod
    def write(cls, filepath: Path, segments: List[Tuple[str, str, str, str]],
              src_lang: str, tgt_lang: str) -> int:
        """
        Записывает XLSX файл с форматированием

        Args:
            filepath: Путь к файлу
            segments: [(src_text, tgt_text, src_lang, tgt_lang), ...]
            src_lang: Исходный язык
            tgt_lang: Целевой язык

        Returns:
            Количество записанных строк
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Translation Memory"

        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Стили для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки
        headers = [
            f"Source ({src_lang})",
            f"Target ({tgt_lang})",
            "Source Language",
            "Target Language"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Записываем данные
        written = 0
        seen = set()
        row = 2

        for src_text, tgt_text, seg_src_lang, seg_tgt_lang in segments:
            # Избегаем дубликатов
            key = (src_text.strip(), tgt_text.strip())
            if key in seen:
                continue
            seen.add(key)

            # Определяем языки
            actual_src = seg_src_lang if seg_src_lang != "unknown" else src_lang
            actual_tgt = seg_tgt_lang if seg_tgt_lang != "unknown" else tgt_lang

            # Записываем строку
            data = [src_text, tgt_text, actual_src, actual_tgt]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border

                # Выравнивание для текста
                if col <= 2:  # Колонки с текстом
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:  # Колонки с языками
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1
            written += 1

        # Автоширина колонок
        cls._auto_adjust_columns(ws)

        # Замораживаем первую строку
        ws.freeze_panes = "A2"

        # Сохраняем
        wb.save(str(filepath))

        logger.info(f"XLSX written: {filepath} ({written} rows)")
        return written

    @classmethod
    def write_simple(cls, filepath: Path, segments: List[Tuple[str, str, str, str]],
                     src_lang: str, tgt_lang: str) -> int:
        """Простая запись без форматирования для больших файлов"""
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Translation Memory")

        # Заголовки
        ws.append([f"Source ({src_lang})", f"Target ({tgt_lang})"])

        written = 0
        seen = set()

        for src_text, tgt_text, seg_src_lang, seg_tgt_lang in segments:
            # Избегаем дубликатов
            key = (src_text.strip(), tgt_text.strip())
            if key in seen:
                continue
            seen.add(key)

            ws.append([src_text, tgt_text])
            written += 1

        wb.save(str(filepath))

        logger.info(f"XLSX written (simple): {filepath} ({written} rows)")
        return written

    @staticmethod
    def _auto_adjust_columns(ws):
        """Автоматически подбирает ширину колонок"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        # Ограничиваем максимальную ширину
                        cell_length = min(len(str(cell.value)), 50)
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Устанавливаем ширину с небольшим запасом
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width